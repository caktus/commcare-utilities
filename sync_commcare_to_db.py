import argparse
import re
from collections import defaultdict

from openpyxl import Workbook, load_workbook

VALID_XML_ENTITY = re.compile(r"\w+")


def extract_property_names(case_summary_file, case_types):
    """Get unique property names that appear in OData feed JSON objects
    This function iterates over each case data item in the feed
    and gathers unique property names.
    Args:
        case_summary_file (str): File path to CommCare app case summary
        case_types (str): Case type to extract from the file
    Returns:
        iterator: Iterator of strings of property names
    """
    wb = load_workbook(filename=case_summary_file)
    # iterate through columns A & B, by row
    rows = zip(*wb["All Case Properties"]["A:B"])
    # read the first (header) row
    header_a, header_b = next(rows)
    assert header_a.value == "case_type"
    assert header_b.value == "case_property"
    # read the remaining rows, filtering out any case types we're not looking for
    # and case properties that are not valid XML entities (they are likely
    # "calculated properties" without a property name in the CommCare app)
    properties_by_type = defaultdict(list)
    for case_type, case_property in rows:
        case_type = case_type.value.strip()
        case_property = case_property.value.strip()
        if case_type in case_types and VALID_XML_ENTITY.fullmatch(case_property):
            properties_by_type[case_type].append(case_property)
    return properties_by_type


def generate_source_target_mappings(source_columns):
    """Generate mapping of source column names to target column names for db

    Args:
        source_columns (list): List of source column names

    Returns:
        list: List of tuples where item[0] is source name, and item[1] is target name
    """
    # For some reason, the same values have different labels in different places in CommCare.
    # this maps the case fields to the OData field names that the user is expecting. This
    # list is taken from running a Case Data API query for a single case and removing any
    # field names that start with "properties" or "xforms". The properties come in via
    # source_columns, and I'm not sure what the xforms are.
    # https://confluence.dimagi.com/display/commcarepublic/Case+Data
    static_case_fields = [
        "case_id",
        "closed",
        "closed_by",
        "date_closed",
        "date_modified",
        "domain",
        "id",
        "indexed_on",
        "indices.parent.case_id",
        "indices.parent.case_type",
        "indices.parent.relationship",
        "opened_by",
        "resource_uri",
        "server_date_modified",
        "server_date_opened",
        "user_id",
    ]
    return [(source_col, source_col) for source_col in static_case_fields] + [
        (f"properties.{source_col}", source_col)
        for source_col in sorted(source_columns)
        if source_col not in static_case_fields
    ]


def make_commcare_export_sync_xl_wb(mapping):
    """Create an Excel workbook in format required for commcare-export script

    NB: This does not save the workbook, and will need to call wb.save() on object
    returned by this function in order to persist.


    Args:
        mapping (dict): Dictionary of lists of tuples of form
            {"case_type": [("source_name", "target_name)]}

    Returns:
        obj: An Openpyxl workbook
    """
    sheet_headers = [
        "Data Source",
        "Filter Name",
        "Filter Value",
        "",
        "Field",
        "Source Field",
    ]
    wb = Workbook()
    for i, (case_type, source_target_mappings) in enumerate(mapping.items()):
        if i == 0:
            ws = wb.active
            ws.title = case_type
        else:
            ws = wb.create_sheet(case_type)
        ws.append(sheet_headers)
        ws["A2"] = "case"
        ws["B2"] = "type"
        ws["C2"] = case_type

        row_offset = 2
        for idx, item in enumerate(source_target_mappings):
            row_num = idx + row_offset
            # NOTE: Columns F, E are not in the order they appear in Excel (E, F)
            ws[f"F{row_num}"], ws[f"E{row_num}"] = item
    return wb


def main(case_summary_file, case_types, output_file_path):
    """Do everything required to export and report on commcare export
    Args:
        case_summary_file (str): File path to CommCare app case summary
        case_types (list): The case types (i.e, "contact", "lab_result, etc.)
        output_file_path (str): Where the workbook with source-column mappings lives
    """
    print(f"Retrieving data from {case_summary_file} and extracting column names")
    properties_by_type = extract_property_names(case_summary_file, case_types)

    mapping = {
        case_type: generate_source_target_mappings(properties)
        for case_type, properties in properties_by_type.items()
    }

    print(f"Generating a temporary Excel workbook to {output_file_path}")
    wb = make_commcare_export_sync_xl_wb(mapping)
    wb.save(output_file_path)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--case-summary-file",
        help="File path to CommCare app case summary",
        dest="case_summary_file",
    )
    parser.add_argument(
        "--case-types",
        help='Space-separated list case types (e.g., "patient lab_result contact investigation")',
        dest="case_types",
        nargs="+",
    )
    parser.add_argument(
        "--db",
        help="The db url string of the db that data will be exported to",
        dest="database_url_string",
    )
    parser.add_argument(
        "--output",
        help="The file path to the Excel query file output that will be created",
        dest="output_file_path",
    )
    args = parser.parse_args()
    main(
        args.case_summary_file, args.case_types, args.output_file_path,
    )
