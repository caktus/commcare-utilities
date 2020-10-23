import argparse
import json
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from cc_utilities.logger import logger

VALID_PROPERTY_NAME = re.compile(r"[\w-]+")
REMOVE_SQL_CHARS = re.compile(r"[\(\:)]")
INVALID_SQL_CHARS = re.compile(r"[^\w]")

# For some reason, the same values have different labels in different places in
# CommCare. This list is taken from running a Case Data API query for a single
# case and removing any field names that start with "properties" or "xforms".
# The properties are app specific and are determined via the Case Summary
# Excel file, and I'm not sure what the xforms are.
# https://confluence.dimagi.com/display/commcarepublic/Case+Data
STATIC_CASE_FIELDS = [
    "case_id",
    "closed",
    "closed_by",
    "date_closed",
    "date_modified",
    "domain",
    "id",
    "indexed_on",
    "indices.parent.case_id",
    "indices.parent.case_type",
    "indices.parent.relationship",
    "opened_by",
    "resource_uri",
    "server_date_modified",
    "server_date_opened",
    "user_id",
]

# The CommCare API documentation also includes several "special case properties"
# which appear to be standard for cases, but do require the "properties." prefix.
# https://confluence.dimagi.com/display/commcarepublic/Case+Data
SPECIAL_CASE_PROPERTIES = [
    "owner_id",
    "case_name",
    "external_id",
    "case_type",
    "date_opened",
]

# List of form data types to exclude from the SQL export. This was discovered
# by trial and error, comparing the results to the form fields shown on the
# "Create Export" interface in CommCare.
EXCLUDE_FORM_DATA_TYPES = [
    "FieldList",
    "Trigger",
    "Group",
    "Repeat",
]


def make_sql_friendly(value, sub="_"):
    """Some CommCare properties and case types include dashes, which make for
    bothersome SQL queries. Remove any non-alphanumeric or underscore
    characters, then return resulting value.
    First, anything matching REMOVE_SQL_CHARS is removed; then, anything
    matching INVALID_SQL_CHARS is replaced with `sub`.
    """
    return INVALID_SQL_CHARS.sub(sub, REMOVE_SQL_CHARS.sub("", value))


def remove_prefix(text, prefix):
    "Removes prefix from text. From: https://stackoverflow.com/a/16891418/166053"
    if text.startswith(prefix):
        return text[len(prefix) :]
    return text


def ws_dict_reader(ws_range):
    "Iterates over an openpyxl worksheet range like a csv DictReader"
    rows = zip(*ws_range)
    headers = next(rows)
    for row in rows:
        yield {header.value: cell.value for header, cell in zip(headers, row)}


def extract_property_names(case_summary_file, case_types):
    """Get unique property names that appear in the Case Summary Excel file.
    This function iterates over each property on the All Case Properties tab
    and gathers unique property names.
    Args:
        case_summary_file (str): File path to CommCare app case summary
        case_types (list): Case types to extract from the file
    Returns:
        iterator: Iterator of strings of property names
    """
    wb = load_workbook(filename=case_summary_file)
    # read the remaining rows, filtering out any case types we're not looking for
    # and case properties that are not valid XML entities (they are likely
    # "calculated properties" without a property name in the CommCare app)
    properties_by_type = defaultdict(list)
    # iterate through columns A & B, by row
    for row in ws_dict_reader(wb["All Case Properties"]["A:B"]):
        case_type = row["case_type"].strip()
        case_property = row["case_property"].strip()
        if case_type in case_types and VALID_PROPERTY_NAME.fullmatch(case_property):
            properties_by_type[case_type].append(case_property)
    return properties_by_type


def generate_form_table_name(form_name, repeats, row):
    table_name = make_sql_friendly(form_name, sub="_")
    if row["repeat"]:
        table_name = "_".join([table_name, repeats[row["repeat"]]])
    while "__" in table_name:
        table_name = table_name.replace("__", "_")
    return table_name


def read_form_definition(module_name, form_name, workbook):
    worksheet_rows = list(ws_dict_reader(workbook[form_name]["A:Z"]))
    # create the canonical list of repeats (will be stashed in their own tables)
    repeats = {
        # {repeat: sql_prefix} mapping
        row["question_id"]: make_sql_friendly(
            remove_prefix(row["question_id"], row["group"])
        )
        for row in worksheet_rows
        if row["type"] == "Repeat"
    }
    result = defaultdict(list)
    for row in worksheet_rows:
        if row["type"] in EXCLUDE_FORM_DATA_TYPES:
            continue
        table_name = generate_form_table_name(form_name, repeats, row)
        result[table_name].append(
            "form."
            + make_sql_friendly(remove_prefix(row["question_id"], "/data/"), sub=".")
        )
    return result


def extract_form_data_fields(form_summary_file):
    """Get unique property names that appear in the Form Summary Excel file.
    This function iterates over each data element on the form definition tabs
    and gathers unique form and form data element names.
    Args:
        form_summary_file (str): File path to CommCare app case summary
    Returns:
        iterator: A mapping of {sql_friendly_form_name: [data_element_1, ...]}
    """
    wb = load_workbook(filename=form_summary_file)
    properties_by_type = {}
    # iterate through columns A & B, by row
    for row in ws_dict_reader(wb["All Forms"]["A:B"]):
        if row["form_name"] in wb:
            properties_by_type.update(
                read_form_definition(row["module_name"], row["form_name"], wb)
            )
    return properties_by_type


def generate_source_target_mappings(source_columns):
    """Generate mapping of source column names to target column names for db. Prepends
    STATIC_CASE_FIELDS and SPECIAL_CASE_PROPERTIES (defined above) to the output.

    Args:
        source_columns (list): List of source column names

    Returns:
        list: List of tuples where item[0] is source name, and item[1] is target name
    """
    property_names = SPECIAL_CASE_PROPERTIES + sorted(
        prop for prop in source_columns if prop not in SPECIAL_CASE_PROPERTIES
    )
    return [(source_col, source_col) for source_col in STATIC_CASE_FIELDS] + [
        # Do not substitute "-" with "_" because in at least once instance, that would
        # result in a duplicate property name ("date_opened").
        (f"properties.{source_col}", make_sql_friendly(source_col, sub=""))
        for source_col in property_names
        if source_col not in STATIC_CASE_FIELDS
    ]


def make_commcare_export_sync_xl_wb(mapping):
    """Create an Excel workbook in format required for commcare-export script

    NB: This does not save the workbook, and will need to call wb.save() on object
    returned by this function in order to persist.


    Args:
        mapping (dict): Dictionary of lists of tuples of form
            {"case_type": [("source_name", "target_name)]}

    Returns:
        obj: An Openpyxl workbook
    """
    sheet_headers = [
        "Data Source",
        "Filter Name",
        "Filter Value",
        "",
        "Field",
        "Source Field",
    ]
    wb = Workbook()
    for i, (data_source, case_type, source_target_mappings) in enumerate(mapping):
        # Some case types have dashes in them, which makes SQL queries more
        # bothersome to write.
        ws_title = make_sql_friendly(case_type, sub="")
        if i == 0:
            ws = wb.active
            ws.title = ws_title
        else:
            ws = wb.create_sheet(ws_title)
        ws.append(sheet_headers)
        ws["A2"] = data_source
        if data_source == "case":
            ws["B2"] = "type"
            ws["C2"] = case_type
        elif data_source == "form":
            # https://confluence.dimagi.com/display/commcarepublic/CommCare+Data+Export+Tool#CommCareDataExportTool-ExampleQueryFile
            ws["B2"] = "xmlns.exact"
            ws["C2"] = "http://openrosa.org/formdesigner/abcdef"
        else:
            raise ValueError(f"Unsupported data source {data_source}")
        row_offset = 2
        for idx, item in enumerate(source_target_mappings):
            row_num = idx + row_offset
            # NOTE: Columns F, E are not in the order they appear in Excel (E, F)
            ws[f"F{row_num}"], ws[f"E{row_num}"] = item
    return wb


def get_previous_mapping(state_dir, case_type):
    """Get the previous state of source-target column mappings from a JSON file
    Args:
        state_dir (str): Path to the directory where JSON files were stored
        case_type (str): This is the commcare case type
    Returns:
        mappings (list): List of tuples of form
            ("source_name", "target_name)
    """
    state_path = Path(state_dir).joinpath(f"{case_type}-column-state.json")
    if state_path.exists():
        with open(state_path) as f:
            state = json.load(f)
            return [
                # json will read these as lists, but they need to be tuples for
                # hashing / comparison purposes later on (this is also how the
                # mapping is generated by get_source_and_target_mapping()).
                tuple(x)
                for x in state["column_mappings"]
            ]
    else:
        # Allow state files to be re-generated if needed.
        return []


def save_column_state(state_dir, case_type, mappings):
    """Save the state of source-target column mappings in a JSON file
    Args:
        state_dir (str): Directory where the state file will be saved
        case_type (str): This is the commcare case type
        mappings (list): List of tuples of form
            ("source_name", "target_name)
    Returns: No return, but saves json file to save_path
    """
    state_dir = Path(state_dir)
    state_dir.mkdir(parents=True, exist_ok=True)
    state_path = state_dir.joinpath(f"{case_type}-column-state.json")
    state = {
        "case_type": case_type,
        "column_mappings": mappings,
        "as_of": datetime.now().strftime("%Y_%m_%d-%H_%M_%S"),
    }
    with open(state_path, "w") as f:
        json.dump(state, f, sort_keys=True, indent=2)
        # Add missing newline at end of file.
        f.write("\n")


def main_with_args(
    form_summary_file, case_summary_file, case_types, output_file_path, state_dir
):
    """Do everything required to export and report on commcare export
    Args:
        form_summary_file (str): File path to CommCare app form summary
        case_summary_file (str): File path to CommCare app case summary
        case_types (list): The case types (i.e, "contact", "lab_result, etc.)
        output_file_path (str): Where the workbook with source-column mappings lives
    """
    new_mappings = []
    if case_summary_file:
        logger.info(
            f"Retrieving data from {case_summary_file} and extracting column names"
        )
        properties_by_type = extract_property_names(case_summary_file, case_types)
        new_mappings.extend(
            ("case", case_type, generate_source_target_mappings(properties))
            for case_type, properties in properties_by_type.items()
        )
    if form_summary_file:
        logger.info(
            f"Retrieving data from {form_summary_file} and extracting column names"
        )
        form_data_fields = extract_form_data_fields(form_summary_file)
        new_mappings.extend(
            ("form", table_name, [(f, f) for f in fields])
            for table_name, fields in form_data_fields.items()
        )
    if state_dir:
        for case_type, new_mapping in new_mappings.items():
            previous_mapping = get_previous_mapping(state_dir, case_type)
            if set(previous_mapping) != set(new_mapping):
                save_column_state(state_dir, case_type, new_mapping)

    logger.info(f"Generating a temporary Excel workbook to {output_file_path}")
    wb = make_commcare_export_sync_xl_wb(new_mappings)
    wb.save(output_file_path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--form-summary-file",
        help="File path to CommCare app form summary",
        dest="form_summary_file",
    )
    parser.add_argument(
        "--case-summary-file",
        help="File path to CommCare app case summary",
        dest="case_summary_file",
    )
    parser.add_argument(
        "--case-types",
        help='Space-separated list case types (e.g., "patient lab_result contact investigation")',
        dest="case_types",
        nargs="+",
    )
    parser.add_argument(
        "--state-dir",
        help="The directory where a JSON representation of the column state should "
        "be read/saved for each case and/or form type (optional)",
        dest="state_dir",
    )
    parser.add_argument(
        "--output",
        help="The file path to the Excel query file output that will be created",
        dest="output_file_path",
    )
    args = parser.parse_args()
    main_with_args(
        args.form_summary_file,
        args.case_summary_file,
        args.case_types,
        args.output_file_path,
        args.state_dir,
    )
