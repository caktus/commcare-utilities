import subprocess

from openpyxl import Workbook


def make_commcare_export_sync_xl_wb(source_target_mappings, case_type):
    """Create an Excel workbook in format required for commcare-export script

    NB: This does not save the workbook, and will need to call wb.save() on object
    returned by this function in order to persist.


    Args:
        source_target_mappings (dict): dict of source:target mappings from source field
            to sql db column name
        case_type (str): This is the case type, and gets added as the "Filter value"
            in the workbook, as well as the worksheet name.

    Returns:
        obj: An Openpyxl workbook
    """
    sheet_headers = [
        "Data Source",
        "Filter Name",
        "Filter Value",
        "",
        "Field",
        "Source Field",
        "Alternate Source Field 1",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = case_type
    ws.append(sheet_headers)
    ws["A2"] = "case"
    ws["B2"] = "type"
    ws["C2"] = case_type

    row_offset = 2

    for idx, item in enumerate(
        [(k, source_target_mappings[k]) for k in source_target_mappings]
    ):
        row_num = idx + row_offset
        ws[f"F{row_num}"], ws[f"E{row_num}"] = item

    return wb


def do_commcare_export_to_db(
    database_url_string,
    commcare_project_name,
    wb_file_path,
    commcare_user_name,
    commcare_api_key,
):
    """Run `commcare-export` as subprocess to export data to SQL db

    Args:
        database_url_string (str): Full db url to export to
        commcare_project_name (str): The Commcare project being exported from
        wb_file_path (str): Where the workbook with source-column mappings lives
        commcare_user_name (str): The Commcare username (email address)
        commcare_api_key (str): A Commcare API key for the user
    """
    commands = (
        f"commcare-export --output-format sql "
        f"--output {database_url_string} --project {commcare_project_name} "
        f"--query {wb_file_path} --username {commcare_user_name} "
        f"--auth-mode apikey --password {commcare_api_key}"
    ).split(" ")
    subprocess.run(commands)
