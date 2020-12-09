import re
import time
from urllib.parse import urljoin

import pandas as pd
import requests
from openpyxl import Workbook
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry

from .constants import (
    APPLICATION_STRUCTURE_DEFAULT_TIMEOUT,
    APPLICATION_STRUCTURE_URL,
    BULK_UPLOAD_URL,
    COMMCARE_UPLOAD_STATES,
    LIST_CASES_URL,
)
from .logger import logger


class CommCareUtilitiesError(Exception):
    def __init__(self, message, info):
        super(CommCareUtilitiesError, self).__init__(message)
        self.info = info


def get_application_structure(
    project_slug, cc_username, cc_api_key, app_id, request_timeout=None
):
    """Retrieve data about a CommCare application's structure from the API

    See: https://confluence.dimagi.com/display/commcarepublic/Application+Structure+API

    Args:
        project_slug (str): The name of the CommCare project (aka "domain")
        cc_user_name (str): Valid CommCare username
        cc_api_key (str): Valid CommCare API key
        app_id (str): The id of the application
        request_timeout: Number of seconds for request timeout. This endpoint can take a
            while so the timeout defaults to a large value of 180.

    Returns:
        dict: A dict formed from the JSON returned by the response

    """
    request_timeout = (
        request_timeout if request_timeout else APPLICATION_STRUCTURE_DEFAULT_TIMEOUT
    )
    url = urljoin(APPLICATION_STRUCTURE_URL.format(project_slug), app_id)
    data = dict(format="json",)
    headers = {
        "Authorization": f"ApiKey {cc_username}:{cc_api_key}",
    }
    response = requests.get(url, headers=headers, params=data, timeout=request_timeout)
    if not response.ok:
        message = (
            f"Something went wrong retrieving app structure for app with id "
            f"`{app_id}`.  The response status code was `{response.status_code}`. "
            f"The response text was: `{response.text}`"
        )
        info = {
            "commcare_response_status_code": response.status_code,
            "commcare_response_text": response.text,
        }
        logger.error(message)
        raise CommCareUtilitiesError(message, info)
    return response.json()


def get_commcare_case(
    case_id,
    project_slug,
    cc_username,
    cc_api_key,
    include_child_cases=False,
    include_parent_cases=False,
    request_timeout=30,
):
    url = LIST_CASES_URL.format(project_slug) + case_id
    data = dict(
        child_cases__full=include_child_cases,
        parent_cases__full=include_parent_cases,
        format="json",
    )
    headers = {
        "Authorization": f"ApiKey {cc_username}:{cc_api_key}",
    }
    response = requests.get(url, headers=headers, params=data)
    if not response.ok:
        message = f"Something went wrong retrieving case `{case_id}`"
        info = {
            "commcare_response_status_code": response.status_code,
            "commcare_response_text": response.text,
        }
        logger.error(message)
        raise CommCareUtilitiesError(message, info)
    return response.json()


def get_commcare_cases(
    project_slug,
    cc_username,
    cc_api_key,
    case_type=None,
    include_closed=None,
    owner_id=None,
    user_id=None,
    limit=5000,
    offset=None,
    external_id=None,
    request_timeout=30,
):
    url = LIST_CASES_URL.format(project_slug)
    data = dict(
        owner_id=owner_id,
        user_id=user_id,
        type=case_type,
        closed=include_closed,
        limit=limit,
        offset=offset,
        external_id=external_id,
        format="json",
    )
    data = {key: value for (key, value) in data.items() if value is not None}
    headers = {
        "Authorization": f"ApiKey {cc_username}:{cc_api_key}",
    }
    response = requests.get(url, headers=headers, params=data)
    if not response.ok:
        message = "Something went wrong downloading data from CommCare"
        info = {
            "commcare_response_status_code": response.status_code,
            "commcare_response_text": response.text,
        }
        logger.error(message)
        raise CommCareUtilitiesError(message, info)
    return response.json()["objects"]


def upload_data_to_commcare(
    df,  # can be dict or pd.DataFrame
    project_slug,
    case_type,
    search_column,
    cc_username,
    cc_api_key,
    create_new_cases="on",
    search_field="case_id",
    request_timeout=30,
    file_name_prefix="",
):
    retry_strategy = Retry(
        total=3, backoff_factor=6, status_forcelist=[500], method_whitelist=["POST"]
    )
    headers = {
        "Authorization": f"ApiKey {cc_username}:{cc_api_key}",
    }
    adapter = HTTPAdapter(max_retries=retry_strategy)
    session = requests.Session()
    session.headers.update(headers)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    url = BULK_UPLOAD_URL.format(project_slug)
    if not isinstance(df, pd.DataFrame):
        df = pd.DataFrame(df)
    assert (
        search_column in df.columns
    ), f"Data items must have property '{search_column}'"
    files = {
        "file": (
            f"{file_name_prefix}{case_type}.csv",
            df.to_csv(index=False),
            "application/csv",
            {"charset": "UTF-8"},
        )
    }
    body = dict(
        case_type=case_type,
        search_column=search_column,
        search_field=search_field,
        create_new_cases=create_new_cases,
    )
    req = requests.Request("POST", url, data=body, files=files)
    prepped = session.prepare_request(req)
    response = session.send(prepped, timeout=request_timeout)

    if not response.ok:
        message = (
            f"Something went wrong uploading data to CommCare: "
            f"Status code: {response.status_code} | Reason: {response.reason} "
        )
        info = {
            "commcare_response_reason": response.reason,
            "commcare_response_status_code": response.status_code,
            "commcare_response_text": response.text,
        }
        logger.error(message)
        raise CommCareUtilitiesError(message, info)
    while True:
        seconds = 2
        logger.info(f"Sleeping {seconds} seconds and checking upload status...")
        time.sleep(seconds)
        response_ = session.get(response.json()["status_url"])

        if response_.json()["state"] == COMMCARE_UPLOAD_STATES["failed"]:
            msg = (
                f"Something went wrong uploading data to CommCare. Result state was "
                f"{COMMCARE_UPLOAD_STATES['failed']}"
            )
            logger.error(msg)
            raise CommCareUtilitiesError(msg, None)

        if response_.json()["state"] != COMMCARE_UPLOAD_STATES["success"]:
            logger.info("Upload not ready, will wait")
            continue

        errors = response_.json()["result"]["errors"]
        if (
            response_.json()["state"] == COMMCARE_UPLOAD_STATES["success"]
            and len(errors) == 0
        ):
            logger.info("Succesfully uploaded. All done.")
            break

        if (
            response_.json()["state"] == COMMCARE_UPLOAD_STATES["success"]
            and len(errors) > 0
        ):
            errors_string = ", ".join(
                [f"{error['title']}: {error['description']}" for error in errors]
            )
            msg = f"Something went wrong uploading data to CommCare: {errors_string}"
            logger.error(msg)
            raise CommCareUtilitiesError(msg, errors)


def chunk_list(lst, chunk_size):
    """Yield successive `chunk_size` chunks from `lst`"""
    for i in range(0, len(lst), chunk_size):
        yield lst[i : i + chunk_size]


def make_sql_friendly(value, invalid_sql_chars=re.compile(r"[^\w]")):
    """Some CommCare properties and case types include dashes, which make for
    bothersome SQL queries. Remove any non-alphanumeric or underscore
    characters, then return resulting value."""
    # Do not substitute "-" with "_" because in at least once instance, that would
    # result in a duplicate property name ("date_opened").
    return invalid_sql_chars.sub("", value)


def make_commcare_export_sync_xl_wb(mapping):
    """Create an Excel workbook in format required for commcare-export script

    NB: This does not save the workbook, and will need to call wb.save() on object
    returned by this function in order to persist.


    Args:
        mapping (dict): Dictionary of lists of tuples of form
            {"case_type": [("source_name", "target_name)]}

    Returns:
        obj: An Openpyxl workbook
    """
    sheet_headers = [
        "Data Source",
        "Filter Name",
        "Filter Value",
        "",
        "Field",
        "Source Field",
    ]
    wb = Workbook()
    for idx, (case_type, source_target_mappings) in enumerate(mapping.items()):
        if idx == 0:
            ws = wb.active
        else:
            ws = wb.create_sheet()
        ws.title = case_type
        ws.append(sheet_headers)
        ws["A2"] = "case"
        ws["B2"] = "type"
        ws["C2"] = case_type

        row_offset = 2
        for idx, item in enumerate(sorted(source_target_mappings, key=lambda x: x[1])):
            row_num = idx + row_offset
            # NOTE: Columns F, E are not in the order they appear in Excel (E, F)
            ws[f"F{row_num}"], ws[f"E{row_num}"] = item
    return wb
